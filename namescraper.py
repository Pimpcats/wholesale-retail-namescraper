#!/usr/bin/env python3
"""Wholesale/Retail Name Scraper.

Paste the text you copied from the marketplace page (e.g. Faire). The program
keeps ONLY the store name from each listing -- the first line of every segment --
and drops the noise: the "$X min" lines, star ratings like "4.8", and the
"Free shipping" / "Up to X% off" discount lines. The cleaned names are written
into an Excel (.xlsx) worksheet.

Two ways to use it:

  * GUI (default): just run `python namescraper.py`, paste your text, and click
    "Save to Excel".
  * Command line: `python namescraper.py --clipboard` reads whatever you've
    copied and appends the names to store_names.xlsx, or pass a text file:
    `python namescraper.py listings.txt -o stores.xlsx`.
"""

import argparse
import os
import re
import sys

# --- Patterns for the lines we want to THROW AWAY -------------------------

# "$150 min", "$0 min", "$ 117 min"  (the minimum order line)
PRICE_RE = re.compile(r"^\$\s?[\d,]+\s*min$", re.IGNORECASE)

# A bare star rating on its own line, e.g. "4.8" or "5"
RATING_RE = re.compile(r"^\d+(\.\d+)?$")

# Shipping / discount lines: "Free shipping", "Up to 5% off",
# "Up to $150 off + free shipping", etc.
SHIPPING_RE = re.compile(r"(free shipping|%\s*off|\bup to\b|\$\s?[\d,]+\s*off)",
                         re.IGNORECASE)


def is_noise(line):
    """Return True if a line is a price, rating, or shipping/discount line."""
    return bool(
        PRICE_RE.match(line)
        or RATING_RE.match(line)
        or SHIPPING_RE.search(line)
    )


def extract_names(text):
    """Pull store names out of pasted marketplace text.

    Each listing looks like:

        Store Name
        [4.8]            <- optional rating
        $100 min
        [Free shipping]  <- optional shipping/discount

    The store name is the only line in a listing that is NOT a price, rating, or
    shipping/discount line, so we simply keep those and drop the rest. This works
    whether or not the segments are separated by blank lines.
    """
    names = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        if is_noise(line):
            continue
        names.append(line)
    return names


def write_to_excel(names, path, append=True, skip_duplicates=True):
    """Write store names to column A of an .xlsx file.

    Returns the number of names actually added.
    """
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        sys.exit(
            "The 'openpyxl' package is required.\n"
            "Install it with:  pip install openpyxl"
        )

    existing = []
    if append and os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:
                existing.append(str(row[0]))
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Store Names"
        ws["A1"] = "Store Name"

    seen = {n.lower() for n in existing} if skip_duplicates else set()
    added = 0
    for name in names:
        if skip_duplicates and name.lower() in seen:
            continue
        ws.append([name])
        seen.add(name.lower())
        added += 1

    wb.save(path)
    return added


# --- Command-line mode ----------------------------------------------------

def read_clipboard_text():
    """Best-effort read of the system clipboard as plain text."""
    try:
        import tkinter as tk
        root = tk.Tk()
        root.withdraw()
        text = root.clipboard_get()
        root.destroy()
        return text
    except Exception as exc:  # noqa: BLE001
        sys.exit(f"Could not read the clipboard: {exc}")


def run_cli(args):
    if args.clipboard:
        text = read_clipboard_text()
    elif args.input:
        with open(args.input, "r", encoding="utf-8") as fh:
            text = fh.read()
    else:
        # Read from stdin (e.g. piped text)
        text = sys.stdin.read()

    names = extract_names(text)
    if not names:
        print("No store names found in the text.")
        return

    added = write_to_excel(
        names, args.output, append=not args.overwrite,
        skip_duplicates=not args.allow_duplicates,
    )
    print(f"Found {len(names)} name(s); added {added} new to {args.output}")
    for name in names:
        print(f"  - {name}")


# --- GUI mode -------------------------------------------------------------

def run_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    root = tk.Tk()
    root.title("Wholesale / Retail Name Scraper")
    root.geometry("560x620")

    instructions = (
        "1. Copy the page text from the marketplace.\n"
        "2. Paste it below (or click 'Paste from clipboard').\n"
        "3. Click 'Extract names' to preview, then 'Save to Excel'."
    )
    ttk.Label(root, text=instructions, justify="left").pack(
        anchor="w", padx=10, pady=(10, 5))

    # Paste box
    ttk.Label(root, text="Pasted text:").pack(anchor="w", padx=10)
    paste_box = tk.Text(root, height=12, wrap="word")
    paste_box.pack(fill="both", expand=True, padx=10)

    # Preview box
    ttk.Label(root, text="Extracted store names:").pack(anchor="w", padx=10,
                                                        pady=(8, 0))
    preview = tk.Listbox(root, height=10)
    preview.pack(fill="both", expand=True, padx=10)

    # Options
    opts = ttk.Frame(root)
    opts.pack(fill="x", padx=10, pady=5)
    append_var = tk.BooleanVar(value=True)
    dedupe_var = tk.BooleanVar(value=True)
    ttk.Checkbutton(opts, text="Append to existing file",
                    variable=append_var).pack(side="left")
    ttk.Checkbutton(opts, text="Skip duplicates",
                    variable=dedupe_var).pack(side="left", padx=(15, 0))

    current_names = []

    def do_paste():
        try:
            text = root.clipboard_get()
        except tk.TclError:
            messagebox.showwarning("Clipboard", "Clipboard is empty or not text.")
            return
        paste_box.delete("1.0", "end")
        paste_box.insert("1.0", text)
        do_extract()

    def do_extract():
        nonlocal current_names
        text = paste_box.get("1.0", "end")
        current_names = extract_names(text)
        preview.delete(0, "end")
        for name in current_names:
            preview.insert("end", name)
        count_label.config(text=f"{len(current_names)} name(s) found")

    def do_save():
        if not current_names:
            do_extract()
        if not current_names:
            messagebox.showinfo("Nothing to save", "No store names found.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="store_names.xlsx",
            title="Save store names to Excel",
        )
        if not path:
            return
        try:
            added = write_to_excel(
                current_names, path, append=append_var.get(),
                skip_duplicates=dedupe_var.get(),
            )
        except SystemExit as exc:
            messagebox.showerror("Error", str(exc))
            return
        messagebox.showinfo(
            "Saved", f"Added {added} new name(s) to:\n{path}")

    # Buttons
    btns = ttk.Frame(root)
    btns.pack(fill="x", padx=10, pady=(0, 5))
    ttk.Button(btns, text="Paste from clipboard",
               command=do_paste).pack(side="left")
    ttk.Button(btns, text="Extract names",
               command=do_extract).pack(side="left", padx=5)
    ttk.Button(btns, text="Save to Excel",
               command=do_save).pack(side="left")

    count_label = ttk.Label(root, text="")
    count_label.pack(anchor="w", padx=10, pady=(0, 10))

    root.mainloop()


def main():
    parser = argparse.ArgumentParser(
        description="Extract store names from copied marketplace text into Excel.")
    parser.add_argument("input", nargs="?",
                        help="Text file to read. Omit to use the GUI.")
    parser.add_argument("-o", "--output", default="store_names.xlsx",
                        help="Excel file to write (default: store_names.xlsx)")
    parser.add_argument("--clipboard", action="store_true",
                        help="Read the text from the system clipboard.")
    parser.add_argument("--overwrite", action="store_true",
                        help="Overwrite the Excel file instead of appending.")
    parser.add_argument("--allow-duplicates", action="store_true",
                        help="Keep duplicate names instead of skipping them.")
    parser.add_argument("--gui", action="store_true",
                        help="Force the GUI even if other args are given.")
    args = parser.parse_args()

    if args.gui or (not args.input and not args.clipboard and sys.stdin.isatty()):
        run_gui()
    else:
        run_cli(args)


if __name__ == "__main__":
    main()
